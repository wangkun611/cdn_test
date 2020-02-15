import unittest
from cdn_model_test import FunctionTestCase
import cdn_model_test
import argparse
import os
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import units

parser = argparse.ArgumentParser(description='Test CDN backup node')
parser.add_argument("--ip", help="CDN backup node ip", required=True)
parser.add_argument("--cookies", help="Set http reqest cookies", action='append')
parser.add_argument("domain", help="Domain XLS file", nargs='*')

wbs = []
success = PatternFill('solid', fgColor='FF00FF00')

def suiteFile(file, cookies, suite):
    '''
    根据Excel文件生成TestCase，默认所有Case都成功，设置成绿色背景
    '''
    wb = load_workbook(file)
    wbs.append((wb, file))
    for ws in wb.worksheets:
        for row in ws.iter_rows(2):
            ncols = len(row)
            domain = row[0].value
            https = False
            url = ''
            if ncols > 1:
                https = row[1].value
            if ncols > 2:
                url = row[2].value
            suite.addTest(FunctionTestCase(domain, url=url, methodName='test_http', cell=row[2], cookies=cookies))
            row[2].fill = success
            if https:
                suite.addTest(FunctionTestCase(domain, url=url, methodName='test_https', cell=row[1], cookies=cookies))
                row[1].fill = success
            for ix in range(3, ncols):
                v = row[ix].value
                if type(v) == type('') and len(v) > 0:
                    row[ix].fill = success
                    suite.addTest(FunctionTestCase(domain, url=row[ix].value, methodName='test_http', cell=row[ix], cookies=cookies))

def suite(files, cookies):
    suite = unittest.TestSuite()
    for file in files:
        suiteFile(file, cookies, suite)
    return suite

if __name__ == '__main__':
    args = parser.parse_args()
    cdn_model_test.backupcdn_ip = args.ip
    cookies={}
    for cookie in args.cookies:
        for kv in cookie.split(';'):
            name,value=kv.strip().split('=',1)
            cookies[name] = value

    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite(args.domain, cookies))

    skip = NamedStyle(name="skip")
    skip.fill = PatternFill('solid', fgColor='FFFFDC4C')
    error = NamedStyle(name="error")
    error.fill = PatternFill('solid', fgColor='FFFF0000')

    for (case, desc) in result.errors[0:] + result.failures[0:]:
        case._cell.style = error
        case._cell.comment = Comment(desc, 'AutoTestor', width=units.points_to_pixels(500), height=units.points_to_pixels(100))
    for (case, desc) in result.skipped:
        case._cell.style = skip
        case._cell.comment = Comment(desc, 'AutoTestor')
    for (wb, file) in wbs:
        t = os.path.splitext(file)
        wb.save(''.join((t[0], '_result', t[1])))
